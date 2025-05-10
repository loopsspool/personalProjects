import re
import xlsxwriter
import os

from openpyxl import load_workbook
from app_globals import PARENT_DIR

# TODO: Run file checklist at end of each scrape? update when image found?
# TODO: Make sure run new file checklist if new pokes added to poke_info sheet


def normalize_empty_in_sheet(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == ""):
                cell.value = None


def load_sheet_from_excel(wb_path, sheet_index=0):
    workbook = load_workbook(filename = wb_path, data_only=True)
    sheet = workbook.worksheets[sheet_index]
    normalize_empty_in_sheet(sheet)
    workbook.close()
    return sheet


def cell_value(sheet, row, col):
    return (sheet.cell(row, col).value)


def isnt_empty(sheet, row, col):
    return (cell_value(sheet, row, col) != None)


def is_empty(sheet, row, col):
    return (cell_value(sheet, row, col) == None)


# Returns column number from column name
def get_col_number(sheet, col_name):
    for col in range(1, sheet.max_column):
        if (cell_value(sheet, 1, col) == col_name):
            return col


# Returns column name from column number
def get_col_name(sheet, col_number):
    return(cell_value(sheet, 1, col_number))


def get_last_row(sheet):
    for row in reversed(range(1, sheet.max_row + 1)):
        if any(sheet.cell(row, col).value is not None for col in range(1, sheet.max_column + 1)):
            return row
        

def sheet_pretty_print(sheet):
    for row in sheet.iter_rows(values_only=True):
        print("\t".join(str(cell) if cell is not None else "None" for cell in row))


def is_poke_in_game(poke_num, game):
    if isinstance(poke_num, int): poke_num = str(poke_num).zfill(4)
    game_availability_col = get_col_number(pokemon_info_sheet, game)
    for row in range(2, POKE_INFO_LAST_ROW+1):
        if str(cell_value(pokemon_info_sheet, row, poke_info_num_col)) == poke_num:
            return isnt_empty(pokemon_info_sheet, row, game_availability_col)
        

def poke_isnt_in_game(poke_num, game):
    return not is_poke_in_game(poke_num, game)


image_checklist_filepath = os.path.join(PARENT_DIR, 'Pokemon Images Checklist.xlsx')
def create_file_checklist_spreadsheet():
    # This will always create a new file that overrides an existing one
    workbook = xlsxwriter.Workbook(image_checklist_filepath)
    game_sprite_availability_sheet = workbook.add_worksheet('Game Sprites')
    home_sprite_availability_sheet = workbook.add_worksheet('Home Sprites')
    drawn_availability_sheet = workbook.add_worksheet('Drawn')
    home_menu_sprite_availability_sheet = workbook.add_worksheet('Home Menu Imgs')
    game_sprite_availability_sheet.freeze_panes(1, 3)
    home_sprite_availability_sheet.freeze_panes(1, 3)
    drawn_availability_sheet.freeze_panes(1, 3)
    home_menu_sprite_availability_sheet.freeze_panes(1, 3)

    green = '#40D073'
    red = '#C75451'
    grey = '#404040'
    new_poke_top_border_color = '#D9D9D9'
    formats = {
        # New pokes have strong top border
        "new poke info": workbook.add_format({'top_color': new_poke_top_border_color, 'top': 5}),
        "new poke green": workbook.add_format({'top_color': new_poke_top_border_color, 'top': 5, 'left': 1, 'right': 1, 'bottom': 1, 'align': 'center', 'bg_color': green, 'font_color': green}),
        "new poke red": workbook.add_format({'top_color': new_poke_top_border_color, 'top': 5, 'left': 1, 'right': 1, 'bottom': 1, 'align': 'center', 'bg_color': red, 'font_color': red}),
        "new poke grey": workbook.add_format({'top_color': new_poke_top_border_color, 'top': 5, 'left': 1, 'right': 1, 'bottom': 1, 'align': 'center', 'bg_color': grey, 'font_color': grey}),
        "green": workbook.add_format({'border': 1, 'align': 'center', 'bg_color': green, 'font_color': green}),
        "red": workbook.add_format({'border': 1, 'align': 'center', 'bg_color': red, 'font_color': red}),
        "grey": workbook.add_format({'border': 1, 'align': 'center', 'bg_color': grey, 'font_color': grey})
    }

    # Game sprite availability
    write_availability(workbook, game_sprite_availability_sheet, formats, table="Games")
    # Home sprite availability
    write_availability(workbook, home_sprite_availability_sheet, formats, table="home_filenames")
    # Drawn availability
    write_availability(workbook, drawn_availability_sheet, formats, table="drawn_filenames")
    # Home Menu Imgs availability
    write_availability(workbook, home_menu_sprite_availability_sheet, formats, table="home_menu_filenames")



    print("Spreadsheet finished!")
    print("Finalizing spreadsheet close...")
    workbook.close()


def generate_header_row(workbook, worksheet, is_game_sprites=False):
    from db_utils import GAMES
    
    h_format = workbook.add_format({'bold': True, 'align': 'center', 'bg_color': 'gray', 'bottom': 5, 'top': 1, 'left': 1, 'right': 1})
    worksheet.set_row(0, None, h_format)
    worksheet.write(0, 0, "#")
    worksheet.write(0, 1, "Name")
    worksheet.write(0, 2, "Tags")

    if is_game_sprites:
        col_num = 3
        game_cols = {}
        for game in reversed(GAMES):
            game_name = game[0]
            worksheet.write(0, col_num, game_name)
            game_cols[game_name] = col_num
            worksheet.set_column(col_num, col_num, len(str(game_name)) + 1)    # Setting column width, 1 is for padding
            col_num += 1

        return game_cols
    else:
        worksheet.write(0, 3, "Available")


def write_availability(workbook, worksheet, formats, table):
    from db_utils import get_all_game_filenames_info, get_non_game_filename_info, get_poke_name
    
    if table == "Games":
        game_cols = generate_header_row(workbook, worksheet, is_game_sprites=True)
        all_file_info = get_all_game_filenames_info()
    else:
        generate_header_row(workbook, worksheet)
        all_file_info = get_non_game_filename_info(table)

    longest_values = {"num": 4, "name": 0, "tags": 0}

    prev_poke_num = 0
    is_new_poke = True

    # Starting at 1 because row 0 is the header row
    for i, (poke_info, files) in enumerate(all_file_info.items(), start=1):
        poke_num = poke_info[0]
        poke_name = get_poke_name(poke_num)
        # Could work pulling poke_sprite_form_id[1] (form) + poke_sprite_form_id[2] (sprite type), but I do seperate shiny from sprite_type sometimes so it wouldn't match the true filename
        if table == "Games":
            # Any game would work here for tags, just pulling it out of the next loop so it isn't rewritten for each game
            tags = get_poke_tags(poke_name, files["SV"]["filename"])
        else:
            tags = get_poke_tags(poke_name, files["filename"])
        longest_values = determine_if_longest_length_value_yet(longest_values, poke_name, tags)
        # Putting a top border on if its a new poke
        if prev_poke_num != poke_num:
            is_new_poke = True
            print(f"\rWriting pokemon #{poke_num} file availability...", end='', flush=True)
            sprite_info_format = formats["new poke info"] if table in ("Games", "home_filenames") else None
        else:
            is_new_poke = False
            sprite_info_format = None
        worksheet.write(i, 0, poke_num, sprite_info_format)
        worksheet.write(i, 1, poke_name, sprite_info_format)
        worksheet.write(i, 2, tags, sprite_info_format)

        if table == "Games":
            for game_name, sprite_data in files.items():
                write_sprite_status_for_game(workbook, worksheet, formats, is_new_poke, i, game_cols, game_name, sprite_data["obtainable"], sprite_data["exists"], sprite_data["has_sub"])
        else:
            # Only allowing is_new_poke formatting (top border line) for home sprites for all the non-game tables
            # It's the only one with enough imgs for each poke to be useful, looks cluttered for the other img types 
            format = determine_format_by_boolean(files["exists"]) if table != "home_filenames" else determine_format_by_boolean(files["exists"], is_new_poke)
            worksheet.write(i, 3, "X" if files["exists"] else "M", formats[format])
        
        prev_poke_num = poke_num

    # Setting column width
    padding = 2
    worksheet.set_column(0, 0, longest_values["num"] + padding)
    worksheet.set_column(1, 1, longest_values["name"] + padding)
    worksheet.set_column(2, 2, longest_values["tags"] + padding)

    # Resetting console line after updates from above
    print('\r' + ' '*45 + '\r', end='')


def determine_if_longest_length_value_yet(longest_values_dict, name, tags):
    name_len = len(str(name))
    tags_len = len(str(tags))
    if name_len > longest_values_dict["name"]: longest_values_dict["name"] = name_len
    if tags_len > longest_values_dict["tags"]: longest_values_dict["tags"] = tags_len
    return longest_values_dict
        

def write_sprite_status_for_game(workbook, worksheet, formats, is_new_poke, row, game_cols, game, obtainable, exists, sub):
    game_col = game_cols[game]
    text = denote_file_status(obtainable, exists, sub)
    format = determine_game_sprite_format(text, is_new_poke)

    worksheet.write(row, game_col, text, formats[format])


def determine_game_sprite_format(text, is_new_poke):
    format = "new poke " if is_new_poke else ""
    if text == "U": format += "grey"
    elif text in ("S", "X"): format += "green"
    elif text == "M": format += "red"
    return format


# is_new_poke = False allows omission of it to not draw the top border for a new poke
def determine_format_by_boolean(is_avail, is_new_poke=False):
    format = "new poke " if is_new_poke else ""
    if is_avail: format += "green"
    if not is_avail: format += "red"
    return format


def denote_file_status(obtainable, exists, sub):
    text = ""
    if not obtainable:
        text = "U"
    else:
        if exists:
            # If image is substituted, denote that
            if sub:
                text = "S"
            else:
                text = "X"
        else:
            text = "M"
    return text


def get_poke_tags(poke_name, filename):
    max_split = 1
    max_split += poke_name.count("-")
    split_filename = filename.split("-", max_split)
    # If theres no tags in the filename, return an empty string
    if len(split_filename) > max_split:
        tags = "-" + split_filename[len(split_filename)-1]
        # Getting rid of file extenstion
        #tags = tags[:-4]
    else:
        tags = ""
    return tags


# TODO: Capitalize constants
# TODO: Verify poke_info_last_row used instead of pokemon_files_sheet.max_row
# Spreadsheet For Pokedex Info
pokemon_info_sheet_path = os.path.join(PARENT_DIR, 'Pokemon Info.xlsx')
pokemon_info_sheet = load_sheet_from_excel(pokemon_info_sheet_path)

pokemon_files = load_workbook(filename = image_checklist_filepath, data_only=True)
pokemon_files_sheet = pokemon_files.worksheets[0]
POKE_INFO_LAST_ROW = get_last_row(pokemon_info_sheet)
poke_info_name_col = get_col_number(pokemon_info_sheet, "Name")
poke_info_num_col = get_col_number(pokemon_info_sheet, "#")
poke_info_gen_col = get_col_number(pokemon_info_sheet, "Gen")
poke_info_f_col = get_col_number(pokemon_info_sheet, "Female Variation")
poke_info_mega_col = get_col_number(pokemon_info_sheet, "Mega")
poke_info_giganta_col = get_col_number(pokemon_info_sheet, "Gigantamax")
poke_info_reg_forms_col = get_col_number(pokemon_info_sheet, "Regional Forms")
poke_info_misc_forms_col = get_col_number(pokemon_info_sheet, "Misc Forms")
poke_info_lgpe_col = get_col_number(pokemon_info_sheet, "LGPE")
poke_info_swsh_col = get_col_number(pokemon_info_sheet, "SwSh")
poke_info_bdsp_col = get_col_number(pokemon_info_sheet, "BDSP")
poke_info_la_col = get_col_number(pokemon_info_sheet, "LA")
poke_info_sv_col = get_col_number(pokemon_info_sheet, "SV")
# TODO: Add drawn and menu sprites
poke_files_num_col = get_col_number(pokemon_files_sheet, "#")
poke_files_name_col = get_col_number(pokemon_files_sheet, "Name")
poke_files_tags_col = get_col_number(pokemon_files_sheet, "Tags")
poke_files_filename_col = get_col_number(pokemon_files_sheet, "Filename")
