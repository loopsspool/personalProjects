import openpyxl     # For reading excel workbook
# Must explicitly state this...
from openpyxl import load_workbook

def cell_value(row, col, sheet):
    return (sheet.cell(row, col).value)

def isnt_empty(row, col, sheet):
    return (cell_value(row, col, sheet) != None)

def is_empty(row, col, sheet):
    return (cell_value(row, col, sheet) == None)

# Returns column number from column name
def get_col_number(col_name, sheet):
    for col in range(1, sheet.max_column):
        if (cell_value(1, col, sheet) == col_name):
            return col

# Returns column name from column number
def get_col_name(col_number, sheet):
    return(cell_value(1, col_number, sheet))

pokemon_files = load_workbook(filename = 'C:\\Users\\ejone\\OneDrive\\Desktop\\Code\\Javascript\\p5\\projects\\Pokeball Pokemon Comparison\\Pokemon File-check.xlsx', data_only=True)
pokemon_files_sheet = pokemon_files.worksheets[0]
filename_col = get_col_number("Filename", pokemon_files_sheet)

gen1_games = ["Yellow", "Red-Green", "Red-Blue"]
gen2_games = ["Silver", "Gold", "Crystal"]
gen3_games = ["Ruby-Sapphire", "FireRed-LeafGreen", "Emerald"]
gen4_games = ["Platinum", "HGSS", "Diamond-Pearl"]
gen5_games = ["BW-B2W2"]
gen6_games = ["XY-ORAS"]
# NOTE: Made SM-USUM Gen7, can change later
gen7_games = ["SM-USUM", "LGPE"]
gen8_games = ["BDSP", "Sword-Shield"]
game_dict = {}


# TODO: Out of curiosity...
    # Get a missing back total
    # And a missing animated total
    # And a missing static front total
# Only doing filename_col up because those are where the actual checks need to be made (missing for certain games)
    # And +1 at the end to be inclusive
for col in range(filename_col + 1, pokemon_files_sheet.max_column + 1):
    col_name = get_col_name(col, pokemon_files_sheet)
    print("Getting ", col_name, "...")
    game_dict[col_name] = 0

    for row in range(2, pokemon_files_sheet.max_row):
        if is_empty(row, col, pokemon_files_sheet):
            game_dict[col_name] += 1

print("\n")
gen_totals = [0] * 8
for k, v in game_dict.items():
    if k in gen1_games:
        gen_totals[0] += v
    if k in gen2_games:
        gen_totals[1] += v
    if k in gen3_games:
        gen_totals[2] += v
    if k in gen4_games:
        gen_totals[3] += v
    if k in gen5_games:
        gen_totals[4] += v
    if k in gen6_games:
        gen_totals[5] += v
    if k in gen7_games:
        gen_totals[6] += v
    if k in gen8_games:
        gen_totals[7] += v

    print(k, " has ", v, " missing images")

print("\n")
total_missing = 0
for i, total in enumerate(gen_totals):
    total_missing += total
    print("Gen", (i + 1), " has ", total, " missing images")

print("\n", total_missing, " total missing images")