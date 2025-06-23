import os
from openpyxl import load_workbook

from app_globals import PARENT_DIR




#|================================================================================================|
#|~~~~~~~~~~~~~~~~~~~~~~~~~~[     GENERAL SPREADSHEET FUNCTIONS     ]~~~~~~~~~~~~~~~~~~~~~~~~~~~~~|
#|================================================================================================|

def normalize_empty_in_sheet(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == ""):
                cell.value = None


def load_sheet_from_excel(wb_path, sheet_name):
    workbook = load_workbook(filename = wb_path, data_only=True)
    sheet = workbook[sheet_name]
    normalize_empty_in_sheet(sheet)
    workbook.close()
    return sheet


def cell_value(sheet, row, col):
    return (sheet.cell(row, col).value)


def isnt_empty(sheet, row, col):
    return (cell_value(sheet, row, col) != None)


def is_empty(sheet, row, col):
    return (cell_value(sheet, row, col) == None)


# Returns column number from column name
def get_col_number(sheet, col_name):
    for col in range(1, sheet.max_column + 1):
        if (cell_value(sheet, 1, col) == col_name):
            return col


# Returns column name from column number
def get_col_name(sheet, col_number):
    return(cell_value(sheet, 1, col_number))


def get_last_row(sheet):
    for row in reversed(range(1, sheet.max_row + 1)):
        if any(sheet.cell(row, col).value is not None for col in range(1, sheet.max_column + 1)):
            return row
        

def sheet_pretty_print(sheet):
    for row in sheet.iter_rows(values_only=True):
        print("\t".join(str(cell) if cell is not None else "None" for cell in row))




#|================================================================================================|
#|~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~[     POKEMON FUNCTIONS     ]~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~|
#|================================================================================================|

def is_poke_in_game(poke_num, game):
    if isinstance(poke_num, int): poke_num = str(poke_num).zfill(4)
    game_availability_col = get_col_number(POKEMON_INFO_SHEET, game)
    for row in range(2, POKE_INFO_LAST_ROW+1):
        if str(cell_value(POKEMON_INFO_SHEET, row, POKE_INFO_NUM_COL)) == poke_num:
            return isnt_empty(POKEMON_INFO_SHEET, row, game_availability_col)
        

def poke_isnt_in_game(poke_num, game):
    return not is_poke_in_game(poke_num, game)




#|================================================================================================|
#|~~~~~~~~~~~~~~~~~~~~~~~~~~[     GLOBAL  POKE INFO CONSTANTS     ]~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~|
#|================================================================================================|

# Spreadsheet for Pokedex Info (pokemon, f/region/mega/giganta variants, megas, species forms, game exclusivity, etc)
POKEMON_INFO_SHEET_PATH = os.path.join(PARENT_DIR, 'Pokemon Info.xlsx')
POKEMON_INFO_SHEET = load_sheet_from_excel(POKEMON_INFO_SHEET_PATH, "Clean Data")

POKE_INFO_LAST_ROW = get_last_row(POKEMON_INFO_SHEET)
POKE_INFO_NAME_COL = get_col_number(POKEMON_INFO_SHEET, "Name")
POKE_INFO_NUM_COL = get_col_number(POKEMON_INFO_SHEET, "#")
POKE_INFO_GEN_COL = get_col_number(POKEMON_INFO_SHEET, "Gen")
POKE_INFO_F_COL = get_col_number(POKEMON_INFO_SHEET, "Female Variation")
POKE_INFO_MEGA_COL = get_col_number(POKEMON_INFO_SHEET, "Mega")
POKE_INFO_GIGANTA_COL = get_col_number(POKEMON_INFO_SHEET, "Gigantamax")
POKE_INFO_REG_FORMS_COL = get_col_number(POKEMON_INFO_SHEET, "Regional Forms")
POKE_INFO_MISC_FORMS_COL = get_col_number(POKEMON_INFO_SHEET, "Misc Forms")

POKE_INFO_COSTUMES_COL = get_col_number(POKEMON_INFO_SHEET, "Costumes")