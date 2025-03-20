from PIL import Image, ImageChops   # For downloading image, cropping, and border checking
import os   # For file moving and checking
import openpyxl     # For reading excel workbook
# NOTE: openpyxl is a 1 based index
import wand.image   # For backup file download

# TODO: Purple background when using on animated pngs
# TODO: Decide if you want this script to update spreadsheet, or to run the file checker again
# TODO: Quality not bad but not the best....
# TODO: Convert this from working from the spreadsheet to working within the directory
    # Check if the animated file exists then the static
    # It's annoying having to update the spreadsheet for this and its kind of overkill

# SPREADSHEET DATA
print("Loading spreadsheet...")
pokemon_info = openpyxl.load_workbook('C:\\Users\\ejone\\OneDrive\\Desktop\\Code\\Javascript\\p5\\projects\\Pokeball Pokemon Comparison\\Pokemon File-check.xlsx')
sheet = pokemon_info.worksheets[0]
n_rows = sheet.max_row
n_cols = sheet.max_column

def cell_value(r, c):
    return (sheet.cell(row = r, column = c).value)

def isnt_empty(row, col):
    return (cell_value(row, col) != None)

def is_empty(row, col):
    return (cell_value(row, col) == None)

# Returns column number from column name
def get_col_number(col_name):
    n_cols = sheet.max_column
    for col in range(1, n_cols+1):
        if (cell_value(1, col) == col_name):
            return col

# Returns generation based on game
def gen_finder_from_game(g):
    if g == "Red-Blue" or g == "Red-Green" or g == "Yellow":
        return("Gen1")
    if g == "Crystal" or g == "Gold" or g == "Silver":
        return("Gen2")
    if g == "Emerald" or g == "FireRed-LeafGreen" or g == "Ruby-Sapphire":
        return("Gen3")
    if g == "Diamond-Pearl" or g == "HGSS" or g == "Platinum":
        return("Gen4")
    if g == "BW-B2W2":
        return("Gen5")
    if g == "XY-ORAS-SM-USUM":
        return("Gen6-7")
    if g == "SM-USUM":
        return("Gen7")
    if g == "Sword-Shield":
        return("Gen8")

# Helper function for is_there_a_border
def find_solid_border_corner(im):
    im = im.convert('RGBA')
    # Checking each corners alpha values
        # Sometimes there's a thin line of transparency
            # Without this, that's interpreted as a "border"
    # Checking top left corner alpha
    if im.getpixel((0, 0))[3] != 0:
        return ((0, 0))
    # Checking top right corner alpha
    if im.getpixel((im.width-1, 0))[3] != 0:
        return ((im.width-1, 0))
    # Checking bottom right corner alpha
    if im.getpixel((im.width-1, im.height-1))[3] != 0:
        return ((im.width-1, im.height-1))
    # Checking bottom left corner alpha
    if im.getpixel((0, im.height-1))[3] != 0:
        return ((0, im.height-1))

    # If all corners are transparent, return -1
    return (-1)

# Checks to see if there's an image around a border
# Courtesy of https://stackoverflow.com/questions/10985550/detect-if-an-image-has-a-border-programmatically-return-boolean
def is_there_a_border(img_name):
    im = Image.open(img_name)
    solid_color_reference_corner = find_solid_border_corner(im)
    # If all four corners are transparent, exit function
    if solid_color_reference_corner == -1:
        return (False)
    
    bg = Image.new(im.mode, im.size, im.getpixel(solid_color_reference_corner))
    diff = ImageChops.difference(im, bg)
    diff = ImageChops.add(diff, diff, 2.0, -100)
    bbox = diff.getbbox()
    # If image has border on any side, print filename
    if bbox != (0,0,im.size[0],im.size[1]):
        return (True)

def remove_border(img_name):
    im = Image.open(img_name)
    solid_color_reference_corner = find_solid_border_corner(im)
    # If all four corners are transparent, exit function
    if solid_color_reference_corner == -1:
        return (False)

    bg = Image.new(im.mode, im.size, im.getpixel(solid_color_reference_corner))
    diff = ImageChops.difference(im, bg)
    diff = ImageChops.add(diff, diff, 2.0, -100)
    bbox = diff.getbbox()
    if bbox:
        im = im.crop(bbox)
        im.save(img_name)


# FILES
game_sprite_path = "C:\\Users\\ejone\\OneDrive\\Desktop\\Code\\Javascript\\p5\\projects\\Pokeball Pokemon Comparison\\Images\\Pokemon\\Game Sprites\\"
files = os.listdir(game_sprite_path)
failed_stills_path = "C:\\Users\\ejone\\Onedrive\\Desktop\\Test\\Failed"
failed_still_files = os.listdir(failed_stills_path)

potentially_incorrect = []
has_border = []
still_generated = []

print("Getting pokemon row ranges...")
# Getting row range of cells pokemon are contained in
pokemon_ranges = {}
# Grabbing the name column from the excel file
name_col = sheet[openpyxl.utils.get_column_letter(get_col_number("Name"))]
# Removes header title 
    # Converts to list because it was a tuple before
name_col = list(name_col)
name_col.pop(0)
# Initializing first values
curr_poke = "Bulbasaur"
# Current start range at 2 because 1 is the header row
curr_poke_start_row_range = 2
curr_poke_end_row_range = -1
# Starting at 2 because header row is at 1
row_i = 2
for name in name_col:
    if name.value != curr_poke:
        # Setting end row
        curr_poke_end_row_range = row_i - 1
        # Adding start & end rows to dictionary
        pokemon_ranges[curr_poke] = (curr_poke_start_row_range, curr_poke_end_row_range)

        # Setting new current poke
        curr_poke = name.value
        curr_poke_start_row_range = row_i
        # Setting to -1 just in case
        curr_poke_end_row_range = -1

    # Setting last pokemon since there won't be a name difference because it's out of rows to iterate
    if row_i == n_rows:
        # Setting end row
        curr_poke_end_row_range = n_rows
        # Adding start & end rows to dictionary
        pokemon_ranges[curr_poke] = (curr_poke_start_row_range, curr_poke_end_row_range)
    # Adding to the row iterator
    row_i += 1

print("Checking files and saving frames...")
# Keeps track of how many photos I can converted from animated to static
# TODO: May be slightly off? Check
img_count = 0
col_i = 1
for pokemon_range in pokemon_ranges.values():
    # Getting the range of rows for the pokemon
    rows = sheet[pokemon_range[0]:pokemon_range[1]]
    # (Re) initializing dictionary for each poke
    named_rows = {}
    # Iterating through the rows to find their names
    for row in rows:
        # Minus one because tuple indices start at 0
            # But columns in openpyxl start at 1
        name = row[get_col_number("Filename") - 1].value
        named_rows[name] = row

    # Checking if there's an animated but not a static image
    for name, row in named_rows.items():
        # Skips animated rows, which will be crossed refrenced in their static counterpart
        if "Animated" in name:
            continue
        else:
            # Get animated counterpart of static tag row
            animated_counterpart_row = named_rows[name + "-Animated"]
            # For finding what game or generation to add
                # Necessary here since it is used in determining if a photo can cover bases for SM-USUM and XY-ORAS
            game = ""
            for i in range(len(row)):
                # Check if there's an animated image and not a static one
                if row[i].value == None and animated_counterpart_row[i].value == "x":
                    img_count += 1
                    # If the previous game (due to assignment) was SM-USUM and found an empty XY-ORAS cell
                        # Which is the current iteration of the loop
                        # Continue, bc the file covers the both of them
                    if game == "XY-ORAS-SM-USUM" and cell_value(1, i+1) == "XY-ORAS":
                        continue
                    game = cell_value(1, i+1)
                    # Checks if XY-ORAS column is empty too, since the games share sprites
                        # And if it's a SM-USUM exclusive, XY-ORAS despite having a black fill also has a "u" character tag, saying its unobtainable
                            # So if it's possible, the cell will be empty and red
                    # Minus one because array indices start at 0
                        # But columns in openpyxl start at 1
                    if game == "SM-USUM" and row[get_col_number("XY-ORAS") - 1].value == None:
                        game = "XY-ORAS-SM-USUM"

                    # Generating filename
                    poke_num = row[get_col_number("#") - 1].value
                    file_tags = row[get_col_number("Tags") - 1].value
                    # If there are no tags (cells empty), replace None with empty string
                    if file_tags == None:
                        file_tags = ""

                    filename = poke_num + " " + row[get_col_number("Name") - 1].value + " "
                    if "Back" in name:
                        filename += gen_finder_from_game(game) + file_tags
                    else:
                        filename += gen_finder_from_game(game) + " " + game + file_tags
                    
                    print(filename)

                    # Actually saving the first frame
                    ############ This method has issues with Emerald-Shiny
                    # Courtesy of https://stackoverflow.com/questions/4904940/python-converting-gif-frames-to-png
                    # Error on some files not having a transparency key?
                    try:
                        im = Image.open(game_sprite_path + filename + "-Animated.gif")
                        transparency = im.info['transparency']
                        im.convert('RGBA')
                        im.save("C:\\Users\\ejone\\OneDrive\\Desktop\\Test\\" + filename + ".png", transparency=transparency)
                        still_generated.append(filename)
                        break
                    except:
                        print("No transparency:", filename)
                        im = Image.open(game_sprite_path + filename + "-Animated.gif")
                        im.convert('RGBA')
                        im.save("C:\\Users\\ejone\\OneDrive\\Desktop\\Test\\" + filename + ".png")
                        still_generated.append(filename)
                        break

                    ############ This method has issues with Crystal and saving 1x1 pixel frames
                    # NOTE: Due to the ubiquity of saving 1x1 images, this method of saving is better suited to attempting to save images that have severely failed with borders/transparency
                    # Courtesy of https://stackoverflow.com/questions/21996710/more-problems-extracting-frames-from-gifs
                    # with wand.image(filename=game_sprite_path + filename + "-Animated.gif") as gif:
                    #     for i, first_frame in enumerate(gif.sequence):
                    #         with wand.image(image=first_frame) as frame:
                    #             frame.format = 'PNG'
                    #             if frame.height < 30:
                    #                 potentially_incorrect.append(filename)
                    #                 break
                    #             # Used to save only unsuccessful attempts at previous saves
                    #             if filename + ".png" in failed_still_files:
                    #                 frame.save(filename = "C:\\Users\\ejone\\OneDrive\\Desktop\\Test\\Failed\\" + filename + ".png")
                    #                 break

print("Checking for borders...")
# Borders are caused by a bug in PIL and partly how gifs are created to save space
    # Appear black if there was no change in the frame
saved_imgs_path = "C:\\Users\\ejone\\Onedrive\\Desktop\\Test"
saved_imgs = os.listdir(saved_imgs_path)
for img in saved_imgs:
    # Skipping failed directory
    if img == "Failed":
        continue
    border = is_there_a_border(saved_imgs_path + "\\" + img)
    if border:
        has_border.append(img)
        # Moves image to failed folder
        new_path = saved_imgs_path + "\\Failed\\" + img
        os.rename(saved_imgs_path + "\\" + img, new_path)
        # Removes border
        remove_border(new_path)

if len(potentially_incorrect) > 0:
    print("\n\n-------------POTENTIALLY INCORRECT-------------")
    for pi in potentially_incorrect:
        print(pi)

if len(has_border) > 0:
    print("\n\n-------------HAD BORDERS REMOVED AND MOVED TO FAILED TO CHECK-------------")
    for bordered in has_border:
        print(bordered)

if len(still_generated) > 0:
    print("\n\n-------------STILLS GENERATED-------------")
    for still in still_generated:
        print(still)
        
print("\n\nDone!")
print("Images added:", img_count)
print("Don't forget to run the file checker again to accomodate for the new images!")
