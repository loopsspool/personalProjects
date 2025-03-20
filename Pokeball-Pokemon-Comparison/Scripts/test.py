import requests # To retrieve webpages
from bs4 import BeautifulSoup   # To parse webpages
import re   # For filtering what images to download
import urllib   # For downloading those images to my computer
import os   # For downloading those images to my computer
from PIL import Image   # For converting URL image data to PIL Image object 
import openpyxl     # For reading excel workbook
# Must explicitly state this...
from openpyxl import load_workbook
import string # To access letters easily without having to type them myself in an array

game_sprite_path = "C:\\Users\\ejone\\OneDrive\\Desktop\\Code\\Javascript\\p5\\projects\\Pokeball Pokemon Comparison\\Images\\Pokemon\\Game Sprites\\"
drawn_path = "C:\\Users\\ejone\\OneDrive\\Desktop\\Code\\Javascript\\p5\\projects\\Pokeball Pokemon Comparison\\Images\\Pokemon\\Menu Sprites\\Gen6\\"
drawn_files = os.listdir(drawn_path)
bigger_gen_8_dl_files = os.listdir(game_sprite_path + "initial_downloads_for_border_removal\\")
# page = requests.get("https://archives.bulbagarden.net/w/index.php?title=Category:Squirtle")
# page_soup = BeautifulSoup(page.content, 'html.parser')

for f in bigger_gen_8_dl_files:
    if "Gen8" not in f:
        print(f)

# nums = []
# for f in files:
#     num = int(f[:3])
#     if num not in nums:
#         nums.append(num)
#     if f[-4:] != ".png":
#         print(f, " not a png")

# for i in range(721):
#     if i not in nums:
#         print(i)

# thumbs = page_soup.find_all('div', 'thumb')
# webpage_starter = "https://archives.bulbagarden.net"
# print (thumbs[42].img["alt"])
# print (webpage_starter + thumbs[42].a.get("href"))

# new_page = requests.get(webpage_starter + thumbs[42].a.get("href"))
# new_page_soup = BeautifulSoup(new_page.content, 'html.parser')

# biggest_link = new_page_soup.find("div", "fullImageLink")
# print(biggest_link.a.get("href"))


# for img in pokemon_imgs:
#     if not "Spr b 7s 001.png" in img.attrs['alt']:
#         continue
#     else:
#         print(get_largest_png(img))

#pokemon_files = load_workbook(filename = 'C:\\Users\\ejone\\OneDrive\\Desktop\\Code\\Javascript\\p5\\projects\\Pokeball Pokemon Comparison\\Pokemon File-check.xlsx', data_only=True)
#pokemon_files_sheet = pokemon_files.worksheets[0]

# def cell_value(row, col, sheet):
#     return (sheet.cell(row, col).value)

# def isnt_empty(row, col, sheet):
#     return (cell_value(row, col, sheet) != None)

# def is_empty(row, col, sheet):
#     return (cell_value(row, col, sheet) == None)

# # Returns column number from column name
# def get_col_number(col_name, sheet):
#     for col in range(1, sheet.max_column):
#         if (cell_value(1, col, sheet) == col_name):
#             return col

# # Returns column name from column number
# def get_col_name(col_number, sheet):
#     return(cell_value(1, col_number, sheet))

# filename_col = get_col_number("Filename", pokemon_files_sheet)

# missing_animated_above_gen5 = 0
# for row in range(2, pokemon_files_sheet.max_row):
#     filename = cell_value(row, filename_col, pokemon_files_sheet)
#     if not "-Animated" in filename:
#             continue
#     print(filename)
#     for col in range(filename_col + 1, pokemon_files_sheet.max_column + 1):
#         if get_col_name(col, pokemon_files_sheet) == "BW-B2W2":
#             break
#         if is_empty(row, col, pokemon_files_sheet):
#             missing_animated_above_gen5 += 1
# print(missing_animated_above_gen5)

# # Vanilla Cream is default, so no letter denoter
# creams = [("Caramel_Swirl", "CaS"), ("Lemon_Cream", "LeC"), ("Matcha_Cream", "MaC"), ("Mint_Cream", "MiC"), ("Rainbow_Swirl", "RaS"), ("Ruby_Cream", "RaC"), ("Ruby_Swirl", "RuS"), ("Salted_Cream", "SaC"), ("Vanilla_Cream", "")]
# # Strawberry Sweet is default, so no letter denoter
# sweets = [("Berry_Sweet", "B"), ("Clover_Sweet", "C"), ("Flower_Sweet", "F"), ("Love_Sweet", "L"), ("Ribbon_Sweet", "R"), ("Star_Sweet", "S"), ("Strawberry_Sweet", "")]
# for f in files:
#     if "Alcremie" in f:
#         bulba_code_form = ""
#         for sweet in sweets:
#             if sweet[0] in f:
#                 # Shiny Alcremie only shows sweet, not cream color
#                 if "-Shiny" in f:
#                     bulba_code_form = sweet[1]
#                     break
#                 else:
#                     # If regular color, find cream
#                     for cream in creams:
#                         if cream[0] in f:
#                             bulba_code_form = cream[1] + sweet[1]
#                             break
#         print(bulba_code_form, "     for     ", f, "\n")
            


# page = requests.get("https://archives.bulbagarden.net/wiki/Category:Pok%C3%A9mon_GO_models")
# soup = BeautifulSoup(page.content, 'html.parser')

# img_link = ""
# caption = ""
# # Finding the filenames for each image
# list_div = soup.find('div', {'class': 'mw-category-generated'})
# for img in list_div.find_all('li'):
#     # Stripping everything after the file extension
#     caption = img.text.split(".png")[0]
#     # Getting rid of all the leading new line characters
#     caption = caption.strip("\n")

#     # Shiny denoter
#     if " s" in caption:
#         # TODO: Denote shininess
#         caption = caption.split(" s")[0]

#     # Female denoter
#     if caption[len(caption) - 1] == "f":
#         # TODO: Denote female variation
#         caption = caption[:len(caption) - 2]

#     # Mega denoter
#     if caption[len(caption) - 1] == "M":
#         # TODO: Denote Mega variation
#         caption = caption[:len(caption) - 2]

#     # Regional denoters
#     if caption[len(caption) - 1] == "A":
#         # TODO: Denote Alolan variation
#         caption = caption[:len(caption) - 2]
#     if caption[len(caption) - 1] == "G":
#         # TODO: Denote Galarian variation
#         caption = caption[:len(caption) - 2]

#     if not re.match("GO\d\d\d", caption)

# # TODO: The below is only for Pokemon GO sprites with decorations
#     # So only run this if the filename doesn't follow the format
#         # GO001 -- Default, GO then pokedex #
#         # GO001 s -- Shiny default
# # Gets all table cells (where description of GO variation is)
# lines = soup.find_all('td')
# for line in lines:
#     # Finds specific line of text describing image
#     if "Model of" in line.text:
#         # Splits line based off where description is (first parenthesis)
#         form = line.text.split("(")[1]
#         # Commas are in shiny forms, this eliminates that
#         if "," in form:
#             form = form.split(',')[0]
#         # Otherwise, just gets text within parenthesis
#         else:
#             form = form.split(')')[0]
#         print(form)
