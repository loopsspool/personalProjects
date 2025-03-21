import string # To access letters easily without having to type them myself in an array
from openpyxl import load_workbook
from spreadsheet_funcs import get_col_number

# Pokemon object
class Pokemon:
    def __init__(self, name, number, gen, has_f_var, has_mega, has_giganta, reg_forms, has_type_forms, has_misc_forms, is_in_gen8):
        self.name = name
        self.number = number
        self.gen = gen
        self.has_f_var = has_f_var
        self.has_mega = has_mega
        self.has_giganta = has_giganta
        self.reg_forms = reg_forms
        self.has_type_forms = has_type_forms
        self.has_misc_forms = has_misc_forms
        self.is_in_gen8 = is_in_gen8
        self.missing_imgs = []
        self.missing_gen1_thru_gen4_back_imgs = []

pokedex = []

# This is so when I get kicked from the server I only have to write once where to pick up
poke_num_start_from = 1
# TODO: Generating pokedex should only go to here too
# TODO: This could be put in a text file, with all missing images too
    # So I don't even have to write this after each crash
        # Each new pokemon would rewrite the line of the file where it picked up
        # And delete missing images that the script downloaded
# This is amount of pokemon to get info from after the starter pokemon
pokemon_to_go_after_start = 0

pokemon_img_urls = []

# These are for back generation differences to be able to loop through and concatonate game name to filename
# TODO: Make this a dict?
# TODO: Where is this used?
# gen1_games = ["Yellow", "Red-Green", "Red-Blue"]
# gen2_games = ["Silver", "Gold", "Crystal"]
# gen3_games = ["Ruby-Sapphire", "FireRed-LeafGreen", "Emerald"]
# gen4_games = ["Platinum", "HGSS", "Diamond-Pearl"]
# gen_1_thru_4_games = [gen1_games, gen2_games, gen3_games, gen4_games]

uppers = list(string.ascii_uppercase)
types = ["Normal", "Fighting", "Flying", "Poison", "Ground", "Rock", "Bug", "Ghost", "Steel", "Fire", "Water", "Grass", "Electric", "Psychic", "Ice", "Dragon", "Dark", "Fairy", "Qmark"]
# Vanilla Cream is default, so no letter denoter
creams = [("Caramel_Swirl", "CaS"), ("Lemon_Cream", "LeC"), ("Matcha_Cream", "MaC"), ("Mint_Cream", "MiC"), ("Rainbow_Swirl", "RaS"), ("Ruby_Cream", "RaC"), ("Ruby_Swirl", "RuS"), ("Salted_Cream", "SaC"), ("Vanilla_Cream", "")]
# Strawberry Sweet is default, so no letter denoter
sweets = [("Berry_Sweet", "B"), ("Clover_Sweet", "C"), ("Flower_Sweet", "F"), ("Love_Sweet", "L"), ("Ribbon_Sweet", "R"), ("Star_Sweet", "S"), ("Strawberry_Sweet", "")]

# Spreadsheet For Pokedex Info
pokemon_info = load_workbook(filename = 'C:\\Users\\ejone\\OneDrive\\Desktop\\Code\\Javascript\\p5\\projects\\Pokeball Pokemon Comparison\\Pokemon Info.xlsx', data_only=True)
pokemon_info_sheet = pokemon_info.worksheets[0]
# Spreadsheet for file tracking (what images I do/dont have)
pokemon_files = load_workbook(filename = 'C:\\Users\\ejone\\OneDrive\\Desktop\\Code\\Javascript\\p5\\projects\\Pokeball Pokemon Comparison\\Pokemon File-check.xlsx', data_only=True)
pokemon_files_sheet = pokemon_files.worksheets[0]

save_path_starter = "C:\\Users\\ejone\\OneDrive\\Desktop\\Code\\Javascript\\p5\\projects\\Pokeball Pokemon Comparison\\Images\\Pokemon"
game_save_path = save_path_starter + "\\Game Sprites\\"
gen6_menu_sprite_save_path = save_path_starter + "\\Menu Sprites\\Gen6\\"
gen8_menu_sprite_save_path = save_path_starter + "\\Menu Sprites\\Gen8\\"
drawn_save_path = save_path_starter + "\\Drawn\\"
# TODO: This didn't download gen1-3 female back sprites (Are they different?)
gen1_thru_4_backs_save_path = game_save_path + "\\back_imgs_to_be_filtered\\"
